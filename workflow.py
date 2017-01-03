# -*- coding: utf-8 -*-
"""
Created on Dec 27 2016 by Nathaniel Mahowald

Contributors: Johnny Wong, Jacky Zhu
"""
# This file will contain the developing workflow of the whole project
# as an updating combination of functionality developed elsewhere


import openpyxl #library for pulling data from excel files
import csv #library for outputting csv files
import ssl #deals with SSL Certificate Errors
import os
import pandas as pd
import numpy as np
from urllib.request import Request, urlopen #Modules used to access websites with their URLs
from urllib.error import URLError, HTTPError #Modules used to deal with errors with accessing websites
from http.client import IncompleteRead #Strange error catch
from ssl import CertificateError #Strange error catch and bypass
import time
import lxml.html
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import re
import collections

from lxml import etree # Catching errors with parsing
from bs4 import UnicodeDammit, BeautifulSoup # for HTML parsing
from datetime import datetime

startTime = datetime.now()

import timeout_decorator


class Co:
    """Short for COmpany, this class will contain all the data we need on each company"""
    def __init__(self, name = None, website = None):
        self.name = name
        self.subsidiary = []
        self.reference = [None, None] # first element contains name (string) of reference, second element contains co object of reference
        if website[0:4] != 'http' and website[0:3] != 'See':
            # appends http heading if not already present and
            # checks if the provider is a subordinate of another provider
            website = 'http://' + website
        elif website[0:3] == 'See':
            website = website[4:]
            website = website.rsplit('(', 1)[0] # some website name had parentheses because of multiple subsidiaries e.g. DealerSocket (2)
            website = website.rsplit(' ', 1)[0]
            self.reference[0] = website
        elif website[0:9] == 'Bought by': # one website had "Bought by" instead of "See"
            self.reference[0] = website[10:]
        self.website = website
        self.content = [] # Contains a set of sublists of the form
                          # [dateAccessed, content]
        self.errors = []  # List of all errors encountered during attempts
                          # on this page

"""imports a correctly formatted spreadhseet as a list of Co objects"""
def excelToCo(inpath = '2015 CloudShare - December Final.xlsx', outpath = None):
    coList = []
    curr_dir = os.getcwd() # gets the current directory
    spreadsheet = openpyxl.load_workbook(curr_dir + os.sep + 'data' + os.sep + inpath) # pulls data directly from excel file
    data = spreadsheet.get_sheet_by_name('Data') # specifies which sheet to pull data from

    for i in range(2, data.max_row):
        # data.max_row sometimes returns more than actual number of rows, so the for loop
        # iterates through empty rows, this if statement accounts for that
        if data.cell(row = i, column = 3).value is not None:
            coList.append(Co(data.cell(row = i, column = 2).value,
                             data.cell(row = i, column = 3).value))

    # goes through each co and assigns co objects as references if needed
    for co in coList:
        if co.reference[0] is not None:
            findReference(co, coList)

    np.save(curr_dir + os.sep + 'data' + os.sep + (outpath if outpath else time.strftime("%d_%m_%Y")), coList)
    return coList

# function is called if "See [company]" or "Bought by [company]" is given in Website column for the current co object
# assigns company's co object to the current co object's reference attribute
# also assigns current co object to company's subsidiary list attribute
def findReference(co, coList):
    for company in coList:
        if co.reference[0] in company.name and company.reference[0] is None:
            co.reference[1] = company
            company.subsidiary.append(co)

"""takes a list of Co objects and attempts to fill out their website content,
   errors encounters, and dates of access using basic html methods
   Ignores duplicates/references"""
def coUpdateHTML(colist, lim = -1, outpath = None, timeout = 10):
    #timeout defaults to 10 seconds
    colist = colist[0:lim]
    @timeout_decorator.timeout(5)
    def getHTML(co):

        if (co.reference[0] == None): # ensures not a reference
            context = ssl._create_unverified_context() #bypasses SSL Certificate Verfication (proabably not a good idea, but it got more sites to work)
            req = Request(co.website, headers = {
                'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
                'Accept-Encoding': 'none',
                'Accept-Language': 'en-US,en;q=0.8',
                'Connection': 'keep-alive'}) # changing the header prevents some webscraper blocking techniques
            try:

                response = urlopen(req, context = context).read() # opens the URL and returns data (in bytes)
            except ConnectionResetError:
                print(co.name, '\'s ', 'didn\'t send data')
                co.errors.append(co.name + ' Server didn\'t send data'+ ' '+ co.website)
            except HTTPError as e:
                print(co.name, '\'s ',  'server couldn\'t fulfill the request.')
                co.errors.append(co.name + ' HTTPError: '+ str(e.code) + ' '+ co.website)
                print('Error code: ', e.code)
            except URLError as e:
                print('We failed to reach a server.')
                co.errors.append(co.name + ' URLError ' + repr(e.reason) + ' '+ co.website)
                print(co.name)
                print('Reason: ', repr(e.reason))
            except CertificateError:
                print("SSL Certificate Error with ", co.name)
                co.errors.append(co.name + ' SSL Certificate Error'+ ' ' + co.website)
            except IncompleteRead as e:
                print("IncompleteRead Error with ", co.name)
                co.errors.append(co.name + ' Incomplete Read Error' + ' ' + co.website)
            else:
                ### do not decode before saving into npy file
                # try:
                #     saveWebPage = UnicodeDammit(
                #         response).unicode_markup  # data (in bytes) from opening URL is decoded into String format
                # except UnicodeDecodeError: # some errors have occurred when decoding characters from non-English languages
                #     print("Unicode Decode Error")
                #     co.errors.append(co.name + ' Decode Error' + ' ' + co.website)
                # except etree.ParserError:
                #     print("ParserError with ", co.name)
                #     co.errors.append(co.name + ' Parser Error ' + co.website)
                # except etree.XMLSyntaxError:
                #     print("XMLSyntaxError with ", co.name)
                #     co.errors.append(co.name + ' XML Syntax Error ' + co.website)
                # else:

                """Store the website content"""
                co.content.append([time.strftime("%d/%m/%Y"), response])
                print (co.name + ' is working fine')
                return co

    for co in colist: # iterates through every Co object
        try:
            getHTML(co)
        except timeout_decorator.timeout_decorator.TimeoutError:
            print (co.name + ' has a TimeoutError')
            co.errors.append(co.name + ' TimeoutError' + ' ' + co.website)
    curr_dir = os.getcwd() # gets the current directory
    np.save(curr_dir + os.sep + 'data' + os.sep + (outpath if outpath else time.strftime("%d_%m_%Y")), colist)
                                            # Updates the saved binaries
                                            # if no name is provided it uses the date
    return colist

"""takes a list of Co objects and attempts to fill out the website content,
   of the cos that encountered errors before using selenium web driver
   Ignores duplicates/references"""
def coUpdateSelenium(colist, lim = -1, outpath = None):
    count = 0
    colist = colist[0:lim]
    driver = webdriver.Firefox()
    for co in colist: # iterates through every Co object that errored
        if co.errors != []:

            driver.get(co.website)
            if driver.page_source:
                count = count + 1
                co.content.append([time.strftime("%d/%m/%Y"), driver.page_source])
                print (co.name + ' is working fine with selenium')
            else:
                print ("Selenium issue")
                co.errors.append("Selenium issue")

    driver.close()
    print(str(count))
    curr_dir = os.getcwd() # gets the current directory
    np.save(curr_dir + os.sep + 'data' + os.sep + (outpath if outpath else time.strftime("%d_%m_%Y")), colist)
                                            # Updates the saved binaries
                                            # if no name is provided it uses the date
    return colist



def word_scrape_list_of_sites(colist):

    def combine_counters(list_of_counters):

        total_index = collections.Counter()

        for counter in list_of_counters:
            total_index.update(counter)

        return total_index

    def word_index(co):

        # Words we don't care about
        useless_words = ['the', 'contact', 'us', 'and', 'subscribe', 'to', 'on', 'a', 'our', 'visit', 'all', 'rights', 'reserved',
                    'your', 'for', 'more', 'read', 'their', 'with', 'every', 'you', 'what', 'why', 'how', 'new', 'are', 'that', 'are', 'from', 'can', 'get', 'this',
                    'has', 'will']

        # ignore words shorter than 3 letters, numbers, and words in useless_words
        def validate_word(word):
            if len(word) > 2 and not word[0].isdigit() and word not in useless_words:
                return word

    # Function to scrape text
        def text_scraper(co_html):

            soup = BeautifulSoup(co_html,"html.parser") # get html from site

            # kill all script and style elements
            for script in soup(["script", "style"]):
                script.extract()    # rip it out

            # get text
            text = soup.get_text()

            # break into lines and remove leading and trailing space on each
            lines = (line.strip() for line in text.splitlines())

            # break multi-headlines into a line each
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))

            # drop blank lines
            text = '\n'.join(chunk for chunk in chunks if chunk)

            return text

        word_indexing = collections.Counter() # sets up a counter where we put words into
        words_index = re.findall(r'\w+', text_scraper(co.content[-1][1]).lower()) #find only words and make them lower case

        total_words = len(words_index) # find total words on webpage

        # gets rid of words we don't want
        word_in_text = [validate_word(word) for word in words_index if validate_word(word) is not None]

        for word in word_in_text:
            word_indexing[word] += 1
        for word in word_indexing:
            word_indexing[word] = word_indexing[word]/total_words
        return word_indexing

    list_index = combine_counters([word_index(co) for co in colist if co.content != []])
    return list_index # returns a dictionary with words as keys and word densities as values




def npyImport(name = 'binaries.npy'):
    curr_dir = os.getcwd() # gets the current directory
    return np.load(curr_dir + os.sep + 'data' + os.sep + name)


def createCSVFile(filename, rows):
    curr_dir = os.getcwd()

    # adds date and time to end of file name to avoid overwriting
    savepath = os.path.join(curr_dir + "/data/" + filename + time.strftime("%Y-%m-%d(%H_%M_%S)", time.gmtime()) + ".csv")

    with open(savepath, "w", encoding = 'utf-8') as output:
        writer = csv.writer(output, lineterminator='\n')
        writer.writerows(rows)


"""Here is the main workflow"""
# coList = npyImport("01_01_2017.npy") # npy file too big to put on github
# listIndexes = word_scrape_list_of_sites(coList)
# rows = listIndexes.items() # converts dictionary into a list of tuples
# createCSVFile("Words", rows)

r = excelToCo()
r = coUpdateHTML(r, lim = -1)
print (datetime.now() - startTime)






