from urllib.request import Request, urlopen  # Modules used to access websites with their URLs
from urllib.error import URLError, HTTPError  # Modules used to deal with errors with accessing websites
from http.client import IncompleteRead, BadStatusLine  #
from ssl import CertificateError  #
from lxml import etree
from time import gmtime, strftime
from bs4 import UnicodeDammit
import openpyxl  # library for pulling data from excel files
import csv  # library for outputting csv files
import ssl  # deals with SSL Certificate Error
import lxml.html
import os
import sys


# spreadsheet = openpyxl.load_workbook('Test.xlsx')
# data = spreadsheet.get_sheet_by_name('Sheet1')

spreadsheet = openpyxl.load_workbook('2015 CloudShare - December Final.xlsx')  # pulls data directly from excel file
data = spreadsheet.get_sheet_by_name('Data')  # specifies which sheet to pull data from


# checks if a text contains a specified character
def checkForCharacter(character, text):
    for x in text:
        if x == character:
            return True
    return False


# in a specified text, replaces all instances of a specified
# character with a new character
def replaceAll(characterToBeReplaced, newCharacter, text):
    text = list(text)
    for i in range(len(text)):
        if text[i] == characterToBeReplaced:
            text[i] = newCharacter
    return ''.join(text)

savePath = os.getcwd() + '/Webpages ' + strftime("%Y-%m-%d %H_%M_%S", gmtime())

#probably not necessary if we're storing these websites in binary
def checkDirectoryExistence(): #checks if directory for webpages already exists to prevent unintentional overwriting
    return os.path.exists(os.path.join(savePath))

if checkDirectoryExistence():
    folderOverwrite = input("The folder Webpages already exists.\nWould you like to overwrite it? (Y/N): ")
    if not (folderOverwrite == "yes" or folderOverwrite == "Y" or folderOverwrite == "y" or folderOverwrite == "Yes"):
        sys.exit("Program terminated. Please rename existing folder Webpages.") #end program if user does not intend to overwrite

os.makedirs(savePath) #creates a new directory to store html files


class Website:
    def __init__(self, URL, provider, i, depth):
        self.URL = URL
        self.provider = provider
        self.row = i
        self.depth = depth


websites = []  # stores all the URLs from the excel file
for i in range(2, data.max_row):
    URL = data.cell(row=i, column=3).value
    if URL != None: #case for when no website is given
        if URL[0:4] != 'http' and URL[0:3] != 'See':  # appends http heading if not already present and checks if the provider is a subordinate of another provider
            URL = 'http://' + URL
    else:
        URL = ' '
    website = Website(URL, data.cell(row=i, column=2).value, i, 0)
    websites.append(website)

hdr = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
    'Accept-Encoding': 'none',
    'Accept-Language': 'en-US,en;q=0.8',
    'Connection': 'keep-alive'}


WorkingArray = []
WorkingArray.append(("Row", "Provider", "Website"))

ErrorArray = []  # stores all the websites that return an error (later written into a CSV file, each element is a row in the CSV file)
ErrorArray.append(("Row", "Provider", "Error Message", "Error Code", "Website"))

NewWebsites = []
NewWebsites.append(("Row", "Provider","Websites"))

def findOtherLinks(websites, website, html):
    for link in html.xpath('//a/@href'):  # find all links on webpage and add to new list
        # if link[0:3] == "www":
        #     link = "http://" + link
        # if link[0:4] != 'http':
        #     continue
        NewWebsites.append((website.row, website.provider, link))
        # websites.append(Website(link, website.provider, website.row, 1))
    # return websites

def createHTMLFile(fileName, savePath, webpage):
    if checkForCharacter('/', fileName):
        fileName = replaceAll('/', '_', fileName)
    newFile = open(os.path.join(savePath, fileName + '.html'), 'w')
    newFile.write(webpage)
    newFile.close()

def goToURL(websites, website):
    context = ssl._create_unverified_context()  # bypasses SSL Certificate Verfication (probably not a good idea, but it got more sites to work)
    req = Request(website.URL,
                  headers=hdr)  # changing the header to Mozilla/5.0 prevents some webscraper blocking techniques
    try:

        response = urlopen(req, context=context).read()  # opens the URL

        # Various errors have popped up, so I set up exceptions to catch them.
        # There might be a better way to do this.
    except ConnectionResetError:
        print(website.provider, '\'s ', 'server didn\'t send data')
        ErrorArray.append((website.row, website.provider, 'Server didn\'t send data', ' ',
                           website.URL))
    except HTTPError as e:
        print(website.provider, '\'s ', 'server couldn\'t fulfill the request.')
        ErrorArray.append((website.row, website.provider, 'HTTPError', e.code,
                           website.URL))
        print('Error code: ', e.code)
    except URLError as e:
        print('We failed to reach', website.provider, '\'s ', 'server.')
        ErrorArray.append((website.row, website.provider, 'URLError ', repr(e.reason),
                           website.URL))
        print('Reason: ', e.reason)
    except CertificateError:
        print("SSL Certificate Error with ", website.provider)
        ErrorArray.append((website.row, website.provider, 'SSL Certificate Error', ' ',
                           website.URL))
    except IncompleteRead as e:
        print("IncompleteRead Error with ", website.provider)
        ErrorArray.append((website.row, website.provider, 'Incomplete Read Error', ' ',
                           website.URL))
    except BadStatusLine:
        print("BadStatusLine Error with ", website.provider)
        ErrorArray.append(
            (website.row, website.provider, 'Bad Status Line', ' ', website.URL))
    else:  # if no error occurs
        try:
            saveWebPage = UnicodeDammit(response).unicode_markup # data (in bytes) from opening URL is decoded into String format
            html = lxml.html.fromstring(response)

        except UnicodeDecodeError:  # some errors have occurred when decoding characters from non-English languages
            print("Unicode Decode Error with ", website.provider)
            ErrorArray.append(
                (website.row, website.provider, 'Decode Error', ' ', website.URL))
        except etree.ParserError:
            print("ParserError with ", website.provider)
            ErrorArray.append((website.row, website.provider, "Parser Error", " ", website.URL))
        except etree.XMLSyntaxError:
            print("XMLSyntaxError with ", website.provider)
            ErrorArray.append((website.row, website.provider, "XML Synxtax Error", " ", website.URL))
        else:
            if website.depth == 0:
                # websites = findOtherLinks(websites, website, html)
                findOtherLinks(websites, website, html)

            #create an html file of the website
            WorkingArray.append((website.row, website.provider, website.URL))
            fileName = website.URL
            createHTMLFile(fileName, savePath, saveWebPage)
            print(website.provider, ' is working fine')  # write all data in ErrorArray into a CSV file
    # return websites



for website in websites:
    if website.URL == ' ': #when no website is given, skip the loop
        continue;
    elif (website.URL[0:3] == 'See'):  # checks for subordinate case first and ends current loop if true
        print("Duplicate")
        ErrorArray.append(
            (website.row, website.provider, 'Duplicate', ' ', website.URL))
    else:
        # websites = goToURL(websites, website)
        goToURL(websites, website)



# for i in range(len(websites)):  # iterates through every website
#
#     if websites[i] == ' ': #when no website is given, skip the loop
#         continue;
#     if (websites[i][0:3] == 'See'):  # checks for subordinate case first and ends current loop if true
#         print("Duplicate")
#         ErrorArray.append(
#             (i + 2, data.cell(row=i + 2, column=2).value, 'Duplicate', ' ', data.cell(row=i + 2, column=3).value))
#     else:
#
#         context = ssl._create_unverified_context()  # bypasses SSL Certificate Verfication (proabably not a good idea, but it got more sites to work)
#         req = Request(websites[i],
#                       headers=hdr)  # changing the header prevents some webscraper-blocking techniques
#         try:
#
#             response = urlopen(req, context=context).read()  # opens the URL
#
#
#             # Various errors have popped up, so I set up exceptions to catch them.
#             # Should be a better way to do this.
#         except ConnectionResetError:
#             print(data.cell(row=i + 2, column=2).value, '\'s ', 'server didn\'t send data')
#             ErrorArray.append((i + 2, data.cell(row=i + 2, column=2).value, 'Server didn\'t send data', ' ',
#                                data.cell(row=i + 2, column=3).value))
#         except HTTPError as e:
#             print(data.cell(row=i + 2, column=2).value, '\'s ', 'server couldn\'t fulfill the request.')
#             ErrorArray.append((i + 2, data.cell(row=i + 2, column=2).value, 'HTTPError', e.code,
#                                data.cell(row=i + 2, column=3).value))
#             print('Error code: ', e.code)
#         except URLError as e:
#             print('We failed to reach', data.cell(row=i + 2, column=2).value, '\'s ', 'server.')
#             ErrorArray.append((i + 2, data.cell(row=i + 2, column=2).value, 'URLError ', repr(e.reason),
#                                data.cell(row=i + 2, column=3).value))
#             print('Reason: ', e.reason)
#         except CertificateError:
#             print("SSL Certificate Error with ", data.cell(row=i + 2, column=2).value)
#             ErrorArray.append((i + 2, data.cell(row=i + 2, column=2).value, 'SSL Certificate Error', ' ',
#                                data.cell(row=i + 2, column=3).value))
#         except IncompleteRead as e:
#             print("IncompleteRead Error with ", data.cell(row=i + 2, column=2).value)
#             ErrorArray.append((i + 2, data.cell(row=i + 2, column=2).value, 'Incomplete Read Error', ' ',
#                                data.cell(row=i + 2, column=3).value))
#         except BadStatusLine:
#             print("BadStatusLine Error with ", data.cell(row=i + 2, column=2).value)
#             ErrorArray.append(
#                 (i + 2, data.cell(row=i + 2, column=2).value, 'Bad Status Line', ' ', data.cell(row=i + 2, column=3).value))
#         else:  # if no error occurs
#             try:
#                 html = lxml.html.fromstring(response) #converts data from opening URL into htmlElement type
#                 saveWebPage = response.decode() # data (in bytes) from opening URL is decoded into String format
#             except UnicodeDecodeError:  # some errors have occurred when decoding characters from non-English languages
#                 print("Unicode Decode Error with ", data.cell(row=i + 2, column=2).value)
#                 ErrorArray.append(
#                     (i + 2, data.cell(row=i + 2, column=2).value, 'Decode Error', ' ', data.cell(row=i + 2, column=3).value))
#             else:
#
#                 for link in html.xpath('//a/@href'): #find all links on webpage and add to new list
#                     NewWebsites.append((data.cell(row=i + 2, column=2).value, link))
#
#
#                 #create a text file to store html file of the website
#                 WorkingArray.append((i + 2, data.cell(row=i + 2, column=2).value, data.cell(row=i + 2, column=3).value))
#                 fileName = data.cell(row=i + 2, column=2).value
#                 if checkForCharacter('/', fileName):
#                     fileName = replaceAll('/', '_', fileName)
#                 newFile = open(os.path.join(savePath, fileName + '.html'), 'w')
#                 newFile.write(saveWebPage)
#                 newFile.close()
#                 print(data.cell(row=i + 2, column=2).value, ' is working fine')  # write all data in ErrorArray into a CSV file

with open("Nonworking_Websites" + strftime("%Y-%m-%d %H_%M_%S", gmtime()) + ".csv", "w") as output:
    writer = csv.writer(output, lineterminator='\n')
    writer.writerows(ErrorArray)
with open("Working_Websites" + strftime("%Y-%m-%d %H_%M_%S", gmtime()) + ".csv", "w") as output:
    writer = csv.writer(output, lineterminator='\n')
    writer.writerows(WorkingArray)
with open("Subsites" + strftime("%Y-%m-%d %H_%M_%S", gmtime()) + ".csv", "w") as output:
    writer = csv.writer(output, lineterminator='\n')
    writer.writerows(NewWebsites)