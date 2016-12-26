# -*- coding: utf-8 -*-
"""
Created on Dec 26 2016
"""

#This file will contain a set of functions that are passed relative directory
#paths and will output a python binary to a path


import openpyxl #library for pulling data from excel files
import csv #library for outputting csv files
import ssl #deals with SSL Certificate Errors
import os
import pandas as pd
import numpy as np

class Co:
    """Short for COmpany, this class will contain all the data we need on each company"""
    def __init__(self, name = None, website = None):
        self.name = name
        if website[0:4] != 'http' and website[0:3] != 'See':
            #appends http heading if not already present and
            #checks if the provider is a subordinate of another provider
    	       website = 'http://' + website
        elif website[0:3] == 'See':
            True
            #Reference to other Co needed?
        self.website = website
        self.content = None
    def addContent(webByteData):
        self.content = webByteData


"""imports a correctly formatted spreadhseet as a co object"""
def excelToCo(inpath = '2015 CloudShare - December Final.xlsx', outpath = 'binaries'):
    coList = []
    curr_dir = os.getcwd() #gets the current directory
    spreadsheet = openpyxl.load_workbook(curr_dir + os.sep + 'data' + os.sep + inpath) #pulls data directly from excel file
    data = spreadsheet.get_sheet_by_name('Data') #specifies which sheet to pull data from

    for i in range(2, data.max_row):
        coList.append(Co(data.cell(row = i, column = 2).value,
                         data.cell(row = i, column = 3).value))
    np.save(curr_dir + os.sep + 'data' + os.sep + outpath, coList)
    return coList

def coImport(path = 'binaries.bin'):
    return np.load(curr_dir + os.sep + 'data' + os.sep + path)


